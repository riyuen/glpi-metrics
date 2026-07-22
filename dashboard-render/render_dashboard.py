"""
Dashboard renderer for Rundeck-triggered emails.

Given DASHBOARD_ID, loads the app's chrome-free /print/:id route in a headless
browser and writes a screenshot, an Excel workbook, and a self-contained HTML
email body to OUTPUT_DIR. Scheduling, recipients, and actually sending the
email are entirely Rundeck's job (its own cron trigger + email notification
step) — this script only produces the artifacts for that step to use.
"""
import base64
import html
import logging
import os
import re
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from playwright.sync_api import sync_playwright

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

APP_BASE_URL = os.environ.get("APP_BASE_URL", "http://app")
OUTPUT_DIR   = Path(os.environ.get("OUTPUT_DIR", "/output"))

INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def render_dashboard(dashboard_id: str) -> tuple[bytes, dict]:
    with sync_playwright() as p:
        browser = p.chromium.launch()
        try:
            page = browser.new_page(viewport={"width": 1600, "height": 1000})
            page.goto(f"{APP_BASE_URL}/print/{dashboard_id}", wait_until="networkidle", timeout=60000)
            page.wait_for_function("window.__DASHBOARD_EXPORT__?.ready === true", timeout=30000)
            screenshot = page.screenshot(full_page=True)
            export_data = page.evaluate("window.__DASHBOARD_EXPORT__")
        finally:
            browser.close()
    return screenshot, export_data


def sheet_name(title: str, used: set) -> str:
    clean = INVALID_SHEET_CHARS.sub(" ", title or "Widget").strip()[:31] or "Widget"
    name, i = clean, 2
    while name in used:
        suffix = f" ({i})"
        name = clean[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def build_xlsx(widgets: list) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    for w in widgets:
        table = w.get("table")
        if not table:
            continue
        ws = wb.create_sheet(title=sheet_name(w.get("title"), used_names))
        ws.append(table["columns"])
        for row in table["rows"]:
            ws.append(row)
    if not wb.sheetnames:
        wb.create_sheet(title="Vide")

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp_path)
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        os.unlink(tmp_path)


def build_html(dashboard_name: str, screenshot_png: bytes, widgets: list) -> str:
    # Self-contained on purpose (base64-embedded image, not a cid: reference) —
    # no outbound MIME message is assembled here, Rundeck's email step consumes
    # this file directly as the body.
    b64 = base64.b64encode(screenshot_png).decode("ascii")
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'></head><body>",
        f"<h2>{html.escape(dashboard_name or '')}</h2>",
        f"<img src='data:image/png;base64,{b64}' style='max-width:100%;border:1px solid #ddd' />",
    ]
    for w in widgets:
        table = w.get("table")
        if not table:
            continue
        parts.append(f"<h3>{html.escape(w.get('title') or '')}</h3>")
        parts.append("<table border='1' cellspacing='0' cellpadding='4' style='border-collapse:collapse'>")
        parts.append("<tr>" + "".join(f"<th>{html.escape(str(c))}</th>" for c in table["columns"]) + "</tr>")
        for row in table["rows"]:
            parts.append("<tr>" + "".join(f"<td>{html.escape(str(cell))}</td>" for cell in row) + "</tr>")
        parts.append("</table>")
    parts.append("</body></html>")
    return "".join(parts)


def main() -> int:
    dashboard_id = os.environ.get("DASHBOARD_ID")
    if not dashboard_id:
        log.error("DASHBOARD_ID env var is required")
        return 1

    try:
        screenshot, export_data = render_dashboard(dashboard_id)
    except Exception:
        log.exception("Failed to render dashboard %s", dashboard_id)
        return 1

    widgets = export_data.get("widgets", [])
    dashboard_name = export_data.get("dashboardName") or dashboard_id

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "dashboard.png").write_bytes(screenshot)
    (OUTPUT_DIR / "dashboard.xlsx").write_bytes(build_xlsx(widgets))
    (OUTPUT_DIR / "dashboard.html").write_text(build_html(dashboard_name, screenshot, widgets), encoding="utf-8")

    log.info("Rendered dashboard %s (%s) to %s", dashboard_id, dashboard_name, OUTPUT_DIR)
    return 0


if __name__ == "__main__":
    sys.exit(main())
